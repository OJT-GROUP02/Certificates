import openpyxl
import psycopg2
import itertools

from operator import itemgetter
from openpyxl.drawing.image import Image
from datetime import datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, fills, numbers, PatternFill, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.descriptors.excel import UniversalMeasure, Relation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


wb = Workbook()

wb.create_sheet("Page 4", 0)

ws = wb["Page 4"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.50
ws.page_margins.rigt = 0.50
document_font = Font(name="Times New Roman", size=12)

# column sizes
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 9
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 25

# Image
img = Image(r"static/images/bicol-university-logo.png")
img.anchor = 'B1'
img.width = 105
img.height = 105
ws.add_image(img)

# Header
ws['B1'].value = "Republic of the Philippines"
ws.merge_cells('B1:F1')
ws['B2'].value = 'Bicol University'
ws.merge_cells('B2:F2') 
ws['B3'].value = 'OFFICE OF THE UNIVERSITY REGISTRAR'
ws.merge_cells('B3:F3')
ws['B3'].font = Font(name="Times New Roman", size=12, bold=True)
ws['B4'].value = 'Legazpi City'
ws.merge_cells('B4:F4')

ws['B5'].value = 'Tel. No (052) 820-6809'
ws.merge_cells('B5:F5')
ws['B6'].value ='E-mail Add: bu_uro@yahoo.com'
ws.merge_cells('B6:F6')

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        cell.font = document_font

ws['B7'].value = "ISO 9001:2008"
ws['B7'].font = Font(color = "57C4E5")
ws['B7'].alignment = Alignment(horizontal='left')

ws['B8'].value = "Certificate No."
ws['B8'].font = Font(color = "57C4E5")
ws['B8'].alignment = Alignment(horizontal='left')

ws['B9'].value = "TUV100 05 1782"
ws['B9'].font = Font(color = "57C4E5")
ws['B9'].alignment = Alignment(horizontal='left')

ws['A12'].value = "C E R T I F I C A T I O N"
ws.merge_cells('A12:F13')
ws['A12'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['A12'].font = Font(size='23', bold=True)

ws['B15'].value = 'TO WHOM IT MAY CONCERN:'
ws['B15'].font = Font(name="Times New Roman", size=12, bold=True)
ws['A15'].alignment = Alignment(horizontal='left')

# End of Header

# Database Connection

cur = psycopg2.connect(database='certification_db', user='postgres',
                        password='1612', host='localhost',
                        port="5432").cursor()

student_list = []
cur.execute('SELECT * from undergrad_stud')
for student in cur:
    stud_row = list(student)
    student_list.append(stud_row)

course_list = []
cur.execute('SELECT * from degree_courses')
for course in cur:
    course_row = list(course)
    course_list.append(course_row)

major_list = []
cur.execute('SELECT * from majors')
for major in cur:
    major_row = list(major)
    major_list.append(major_row)

award_list = []
cur.execute('SELECT * from awards')
for award in cur:
    award_row = list(award)
    award_list.append(award_row)

reg_list = []
cur.execute('SELECT * from registrar')
for reg in cur:
    reg_row = list(reg)
    reg_list.append(reg_row)

# Body

rows = ws.iter_cols(min_row=16, min_col=1, max_row=40, max_col=9)
for row in rows:
    for cell in row:
        cell.font = document_font

# 1st Paragraph
ws['B18'].value = '         This is to certify that Ms. ' + student_list[2][1].upper() + ' Q. ' + student_list[2][3].upper() + ' has graduated with the degree of'
ws.merge_cells('B18:F18')
ws['B19'].value =  course_list[2][1] + ' (' + course_list[2][2] + ') ' +'major in ' + major_list[2][2] + ', ' + award_list[2][1]
ws.merge_cells('B19:F19')
ws['B20'].value = ' on ' + 'April 03, 1997 per Resolution No. 1, s. 1997 of the Board of Regents, Bicol University.'
ws.merge_cells('B20:F20')

# 2nd Paragraph
ws['B22'].value = '         It is further certified that she is of good moral character and has never been subjected to '
ws.merge_cells('B23:F23')
ws['B23'].value = 'any disciplinary action during her entire stay in this University.'
ws.merge_cells('B24:F24')

# 3rd Paragraph
ws['B25'].value = '         Issued this 21st day of June, 2010 upon the request of Ms. ' + student_list[2][3] + ' for reference purposes.'
ws.merge_cells('B26:F26')

# Registrar
center_align = Alignment(wrap_text=True, horizontal='center', vertical='center')
bold_font = Font(bold=True, name="Times New Roman", size=12)
ws['F31'].value = reg_list[0][1].upper()
ws['F31'].font = bold_font
ws['F31'].alignment = center_align
ws['F32'].value = reg_list[0][2]
ws['F32'].alignment = center_align

# Footer
footer_font = Font(name="Arial Black", size=9)
ws['B37'].value = 'BU-F-UREG-54'
ws['B37'].font = footer_font
ws['F37'].value = 'Revision: 1'
ws['F37'].font = footer_font
ws['B38'].value = 'Effectivity Date: Mar. 9,2011'
ws['B38'].font = footer_font

wb.save('static/Certificate_Page4.xlsx')