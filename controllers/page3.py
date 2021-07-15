import itertools
from datetime import datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))

# db connection
def connect(query):
    cur = psycopg2.connect(database='certification_db', user='postgres',
                           password='1612', host='localhost',
                           port="5432").cursor()
    cur.execute(query)
    desc = cur.description
    column_names = [col[0] for col in desc]
    result = [dict(itertools.zip_longest(column_names, row)) for row in
              cur.fetchall()]
    # result = cur.fetchall()

    return result

# QUERIES

# fetch student data
student_query = "SELECT undergrad_stud.first_name, " \
                "undergrad_stud.middle_name, undergrad_stud.last_name, " \
                "undergrad_stud.gender, undergrad_stud.gwa, undergrad_stud.date_graduated, " \
                "degree_courses.course_name, degree_courses.course_abbrev, " \
                "majors.major FROM undergrad_stud " \
                "LEFT JOIN degree_courses " \
                "ON undergrad_stud.course_id = degree_courses.course_id " \
                "LEFT JOIN majors " \
                "ON undergrad_stud.major_id = majors.major_id " \
                "WHERE undergrad_stud.student_id = 2"

student = connect(student_query)

if student[0]["gender"] == 'F':
    stud_address = "Ms."
else:
    stud_address = "Mr."

if student[0]["middle_name"] is not None:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["middle_name"][0].upper()}. ' \
                       f'{student[0]["last_name"].upper()}'
else:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["last_name"].upper()}'

stud_address_lastname = f'{stud_address} {student[0]["last_name"]}'

# registrar
uni_reg_query = "SELECT * FROM registrar " \
                "WHERE registrar_id = 1"
uni_registrar = connect(uni_reg_query)

# dates
current_day = custom_strftime('{S}', dt.now())
current_month_year = custom_strftime('%B, %Y', dt.now())
date_graduated = custom_strftime('%B %d, %Y', student[0]["date_graduated"])

header_query = "SELECT * FROM campus_college_institute WHERE cci_id = 10"
header = connect(header_query)
college_name = header[0]["cci_name"].title()
college_address = header[0]["address"]
campus= header[0]["campus"]

# PAGE 3  

wb = Workbook()

wb.create_sheet("Page3", 0)

ws = wb["Page3"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.50
ws.page_margins.rigt = 0.50

ws.column_dimensions["I"].width = 15

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:I1')
ws.append(['Bicol University'])
ws.merge_cells('A2:I2')
ws.append(['OFFICE OF THE UNIVERSITY REGISTRAR'])
ws.merge_cells('A3:I3')
ws['A3'].font = Font(bold=True)
ws.append(['Legazpi City'])
ws.merge_cells('A4:I4')
ws.append(['Tel No. (052) 820-6809'])
ws.merge_cells('A5:I5')
ws.append(['E-mail Add: bu_uro@yahoo.com'])
ws.merge_cells('A6:I6')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        cell.font = Font(name='Times New Roman', size=12)

rows = ws.iter_cols(min_row=3, min_col=1, max_row=6, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(color="c1312c", name='Times New Roman', size=12)

ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                     bold=True)

# LOGO
bu_logo = Image(r"static/images/bicol-university-logo.png")

bu_logo.height = 105
bu_logo.width = 105

ws.add_image(bu_logo, "A1")

ws['A7'] = 'ISO 9001:2008'
ws.merge_cells('A7:B7')
ws['A8'] = 'TUV Rheinland ID 910863351'
ws.merge_cells('A8:B8')

rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
for row in rows:
    for cell in row:
        cell.font = Font(color='83afdf', name='Helvetica', size=11, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')

#Title
ws['A11'].value= "C E R T I F I C A T I O N"
ws.merge_cells('A11:I12')

rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(name='Times New Roman', size=20, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')

#Greetings
ws['A15'].value= "TO WHOM IT MAY CONCERN:"
ws.merge_cells('A15:D15')
ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

#Letter Body
ws['A18'].value = f"This is to certify that {student_fullname} has graduated with the degree of"
ws['A18'].alignment = Alignment(horizontal='left', indent=2)
ws.merge_cells('A18:I18')
ws['A19'].value = f"{student[0]['course_name']} ({student[0]['course_abbrev']}), major in {student[0]['major']}"
ws['A19'].alignment = Alignment(horizontal='left')
ws.merge_cells('A19:I19')
ws['A20'].value = f"from {college_name}, {campus}, {college_address} on {date_graduated}"
ws['A20'].alignment = Alignment(horizontal='left')
ws.merge_cells('A20:I20')
ws['A21'].value = f"per Board of Regents Referendum No. 02-A, s. 2015 having a General Weighted Average (GWA)"
ws['A21'].alignment = Alignment(horizontal='left')
ws.merge_cells('A21:I21')
ws['A22'].value = f"of {student[0]['gwa']}"
ws['A22'].alignment = Alignment(horizontal='left')
ws.merge_cells('A22:I22')


ws['A24'].value = f"Issued this {current_day} day of {current_month_year} upon the request of interested party for reference purposes."
ws['A24'].alignment = Alignment(horizontal='left', indent=2)
ws.merge_cells('A24:I24')


font = Font(name='Times New Roman', size=12)
rows = ws.iter_cols(min_row=18, min_col=1, max_row=25, max_col=9)
for row in rows:
    for cell in row:
        cell.font = font

#Signatories
ws['G30'].value = f"{uni_registrar[0]['registrar_name'].upper()}"
ws['G30'].alignment = Alignment(horizontal='center')
ws['G30'].font = Font(bold=True, name='Times New Roman', size=12)
ws.merge_cells('G30:I30')
ws['G31'].value = f"{uni_registrar[0]['registrar_position']}"
ws['G31'].alignment = Alignment(horizontal='center')
ws['G31'].font = Font(name='Times New Roman', size=12)
ws.merge_cells('G31:I31')



# Footer
ws['A40'].value = "BU-F-UREG-05"
ws['A40'].alignment = Alignment(horizontal='left')
ws['A40'].font = Font(bold=True, size=10)
ws.merge_cells('A40:B40')
ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
ws['A41'].alignment = Alignment(horizontal='left')
ws['A41'].font = Font(bold=True, size=10)
ws.merge_cells('A41:C41')

ws['H40'].value = "Revision 1"
ws['H40'].alignment = Alignment(horizontal='right')
ws['H40'].font = Font(bold=True, size=10)
ws.merge_cells('H40:I40')

wb.save('static/Certificate_Page3.xlsx')
