import itertools
from datetime import datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))


# db connection
def connect(query):
    cur = psycopg2.connect(database='newdb', user='postgres',
                           password='april17', host='localhost',
                           port="5432").cursor()
    cur.execute(query)
    desc = cur.description
    column_names = [col[0] for col in desc]
    result = [dict(itertools.zip_longest(column_names, row)) for row in
              cur.fetchall()]
    # result = cur.fetchall()

    return result


# QUERIES

# fetch student data
student_query = "SELECT undergrad_stud.first_name, " \
                "undergrad_stud.middle_name, undergrad_stud.last_name, " \
                "undergrad_stud.gender, undergrad_stud.date_graduated, " \
                "degree_courses.course_name, degree_courses.course_abbrev, " \
                "majors.major FROM undergrad_stud " \
                "LEFT JOIN degree_courses " \
                "ON undergrad_stud.course_id = degree_courses.course_id " \
                "LEFT JOIN majors " \
                "ON undergrad_stud.major_id = majors.major_id " \
                "WHERE undergrad_stud.student_id = 1"

student = connect(student_query)

if student[0]["gender"] == 'F':
    stud_address = "Ms."
else:
    stud_address = "Mr."

if student[0]["middle_name"] is not None:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["middle_name"][0].upper()}. ' \
                       f'{student[0]["last_name"].upper()}'
else:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["last_name"].upper()}'

stud_address_lastname = f'{stud_address} {student[0]["last_name"]}'

# registrar
uni_reg_query = "SELECT * FROM registrar " \
                "WHERE registrar_id = 1"
uni_registrar = connect(uni_reg_query)

# dates
current_day = custom_strftime('{S}', dt.now())
current_month_year = custom_strftime('%B, %Y', dt.now())
date_graduated = custom_strftime('%B %d, %Y', student[0]["date_graduated"])

# WORKBOOK

wb = Workbook()

wb.create_sheet("Page2", 0)

ws = wb["Page2"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.50
ws.page_margins.rigt = 0.50

ws.column_dimensions["I"].width = 14

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:I1')
ws.append(['Bicol University'])
ws.merge_cells('A2:I2')
ws.append(['OFFICE OF THE UNIVERSITY REGISTRAR'])
ws.merge_cells('A3:I3')
ws['A3'].font = Font(bold=True)
ws.append(['Legazpi City'])
ws.merge_cells('A4:I4')
ws.append(['Tel No. (052) 820-6809'])
ws.merge_cells('A5:I5')
ws.append(['E-mail Add: bu_uro@yahoo.com'])
ws.merge_cells('A6:I6')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        cell.font = Font(name='Times New Roman', size=12)

rows = ws.iter_cols(min_row=3, min_col=1, max_row=6, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(color="c1312c", name='Times New Roman', size=12)

ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                     bold=True)

#BU Logo
bu_logo = Image(r"../static/images/bicol-university-logo.png")

bu_logo.height = 105
bu_logo.width = 105

ws.add_image(bu_logo, "A1")

ws['A7'] = 'ISO 9001:2008'
ws.merge_cells('A7:B7')
ws['A8'] = ' Certificate No.'
ws.merge_cells('A8:B8')
ws['A9'] = 'TUV100 05 1782'
ws.merge_cells('A9:B9')

rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
for row in rows:
    for cell in row:
        cell.font = Font(color='83afdf', name='Helvetica', size=11, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')


#Title
ws['A11'].value= "C E R T I F I C A T I O N"
ws.merge_cells('A11:I12')

rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(name='Times New Roman', size=20, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')

#Greetings
ws['A15'].value= "TO WHOM IT MAY CONCERN:"
ws.merge_cells('A15:D15')
ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

#Letter Body
ws['A18'].value = f"This is to certify that {student_fullname} has " \
                  f"graduated with the degree of"
ws['A18'].alignment = Alignment(horizontal='left', indent=5)
ws.merge_cells('A18:I18')
ws['A19'].value = f"{student[0]['course_name']} " \
                  f"({student[0]['course_abbrev']}) " \
                  f"from Bicol University, Legazpi City on " \
                  f"{date_graduated}"
ws['A19'].alignment = Alignment(horizontal='left')
ws.merge_cells('A19:I19')
ws['A20'].value = "per Referendum No. 1, s, 2002 of the Board of Regents."
ws['A20'].alignment = Alignment(horizontal='left')
ws.merge_cells('A20:I20')

ws['A22'].value = "It is further certified that ENGLISH LANGUAGE is the medium of instruction in all"
ws['A22'].alignment = Alignment(horizontal='left', indent=5)
ws.merge_cells('A22:I22')
ws['A23'].value = "courses and in all Program Levels of Bicol University. Thus, all the academic documents such as"
ws['A23'].alignment = Alignment(horizontal='left')
ws.merge_cells('A23:I23')
ws['A24'].value = "Diploma, Transcript of Record, course descriptions, etc. are translated into English language."
ws['A24'].alignment = Alignment(horizontal='left')
ws.merge_cells('A24:I24')

ws['A26'].value = f"Issued this {current_day} day of {current_month_year} " \
                  f"upon the request of " \
                  f"{stud_address_lastname} for reference"
ws['A26'].alignment = Alignment(horizontal='left', indent=5)
ws.merge_cells('A26:I26')
ws['A27'].value = "purposes."
ws['A27'].alignment = Alignment(horizontal='left')
ws.merge_cells('A27:I27')

font = Font(name='Times New Roman', size=12)
rows = ws.iter_cols(min_row=18, min_col=1, max_row=27, max_col=9)
for row in rows:
    for cell in row:
        cell.font = font

#Signatories
ws['F31'].value = f"{uni_registrar[0]['registrar_name'].upper()}"
ws['F31'].alignment = Alignment(horizontal='center')
ws['F31'].font = Font(bold=True, name='Times New Roman', size=12)
ws.merge_cells('F31:I31')
ws['F32'].value = "University Registrar"
ws['F32'].alignment = Alignment(horizontal='center')
ws['F32'].font = Font(name='Times New Roman', size=12)
ws.merge_cells('F32:I32')

# Footer
ws['A40'].value = "BU-F-UREG-04"
ws['A40'].alignment = Alignment(horizontal='left')
ws['A40'].font = Font(bold=True, size=10)
ws.merge_cells('A40:B40')
ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
ws['A41'].alignment = Alignment(horizontal='left')
ws['A41'].font = Font(bold=True, size=10)
ws.merge_cells('A41:C41')

ws['H40'].value = "Revision 1"
ws['H40'].alignment = Alignment(horizontal='right')
ws['H40'].font = Font(bold=True, size=10)
ws.merge_cells('H40:I40')

# wb.save('static/Certificate_Page2.xlsx')
wb.save('D:\\Users\\iveej\\Desktop\\web2py\\applications\\Certificates\\static'
        '\\Certificate_Page2.xlsx')