import itertools
from datetime import datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))



# PAGE 1 WORKBOOK 

wb = Workbook()

wb.create_sheet("Page1", 0)

ws = wb["Page1"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.50
ws.page_margins.rigt = 0.50

ws.column_dimensions["I"].width = 14

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:I1')
ws.append(['Bicol University'])
ws.merge_cells('A2:I2')
ws.append(['GRADUATE SCHOOL'])
ws.merge_cells('A3:I3')
ws['A3'].font = Font(bold=True)
ws.append(['Legazpi City'])
ws.merge_cells('A4:I4')
ws.append(['Tel No. (052) 481-7881'])
ws.merge_cells('A5:I5')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=5, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        cell.font = Font(name='Times New Roman', size=12)

rows = ws.iter_cols(min_row=3, min_col=1, max_row=5, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(color="c1312c", name='Times New Roman', size=12)

ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                     bold=True)

#BU LOGO
bu_logo = Image(r"../static/images/bicol-university-logo.png")

bu_logo.height = 105
bu_logo.width = 105

ws.add_image(bu_logo, "A1")

ws['A7'] = 'ISO 9001:2008'
ws.merge_cells('A7:B7')
ws['A8'] = ' Certificate No.'
ws.merge_cells('A8:B8')
ws['A9'] = 'TUV100 05 1782'
ws.merge_cells('A9:B9')

rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
for row in rows:
    for cell in row:
        cell.font = Font(color='83afdf', name='Helvetica', size=11, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')

#Title
ws['A11'].value= "C E R T I F I C A T I O N"
ws.merge_cells('A11:I12')

rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(name='Times New Roman', size=20, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')

#Greetings
ws['A15'].value= "TO WHOM IT MAY CONCERN:"
ws.merge_cells('A15:D15')
ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

#Letter Body
ws['A18'].value = "On the basis of records on file in this Office, this is to certify that Ms. NOEMI B. LUMBAO,"
ws['A18'].alignment = Alignment(horizontal='left', indent=2)
ws.merge_cells('A18:I18')
ws['A19'].value = "a graduate student of this University, has Completed the Academic Requirements (CAR) for the"
ws['A19'].alignment = Alignment(horizontal='left')
ws.merge_cells('A19:I19')
ws['A20'].value = "degree Master of Arts in Education (MAED), major in Administration and Supervision."
ws['A20'].alignment = Alignment(horizontal='left')
ws.merge_cells('A20:I20')

ws['A22'].value = "Issued this 14th day of June, 2010 upon the request of Ms. Lumbao for reference"
ws['A22'].alignment = Alignment(horizontal='left', indent=2)
ws.merge_cells('A22:I22')
ws['A23'].value = "purposes."
ws['A23'].alignment = Alignment(horizontal='left')
ws.merge_cells('A23:I23')

font = Font(name='Times New Roman', size=12)
rows = ws.iter_cols(min_row=18, min_col=1, max_row=23, max_col=9)
for row in rows:
    for cell in row:
        cell.font = font

#Signatories
ws['F27'].value = "CATHERINE D. LAGATA"
ws['F27'].alignment = Alignment(horizontal='center')
ws['F27'].font = Font(bold=True, name='Times New Roman', size=12)
ws.merge_cells('F27:I27')
ws['F28'].value = "Registrar II"
ws['F28'].alignment = Alignment(horizontal='center')
ws['F28'].font = Font(name='Times New Roman', size=12)
ws.merge_cells('F28:I28')

ws['A30'].value = "Noted:"
ws['A30'].alignment = Alignment(horizontal='left')
ws['A30'].font = Font(name='Times New Roman', size=12)
ws.merge_cells('A30:B30')

ws['A32'].value = "CORAZON N. BAZAR"
ws['A32'].alignment = Alignment(horizontal='center')
ws['A32'].font = Font(bold=True, name='Times New Roman', size=12)
ws.merge_cells('A32:D32')
ws['B33'].value = "University Registrar"
ws['B33'].alignment = Alignment(horizontal='center')
ws['B33'].font = Font(name='Times New Roman', size=12)
ws.merge_cells('B33:C33')

# Footer
ws['A40'].value = "BU-F-UREG-04"
ws['A40'].alignment = Alignment(horizontal='left')
ws['A40'].font = Font(bold=True, size=10)
ws.merge_cells('A40:B40')
ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
ws['A41'].alignment = Alignment(horizontal='left')
ws['A41'].font = Font(bold=True, size=10)
ws.merge_cells('A41:C41')

ws['H40'].value = "Revision 1"
ws['H40'].alignment = Alignment(horizontal='right')
ws['H40'].font = Font(bold=True, size=10)
ws.merge_cells('H40:I40')

wb.save('static/Certificate_Page1.xlsx')