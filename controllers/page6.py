import itertools
from datetime import datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))


# db connection
def connect(query):
    cur = psycopg2.connect(database='certification_db', user='postgres',
                           password='1612', host='localhost',
                           port="5432").cursor()
    cur.execute(query)
    desc = cur.description
    column_names = [col[0] for col in desc]
    result = [dict(itertools.zip_longest(column_names, row)) for row in
              cur.fetchall()]

    return result

# QUERIES

# college
header_query = "SELECT * FROM campus_college_institute WHERE cci_id = 10"
header = connect(header_query)
college_name = header[0]["cci_name"].title()
college_address = header[0]["address"]

# fetch student data
student_query = "SELECT undergrad_stud.first_name, " \
                "undergrad_stud.middle_name, undergrad_stud.last_name, " \
                "undergrad_stud.gender, undergrad_stud.date_graduated, " \
                "degree_courses.course_name, degree_courses.course_abbrev, " \
                "majors.major FROM undergrad_stud " \
                "LEFT JOIN degree_courses " \
                "ON undergrad_stud.course_id = degree_courses.course_id " \
                "LEFT JOIN majors " \
                "ON undergrad_stud.major_id = majors.major_id " \
                "WHERE undergrad_stud.student_id = 4"

student = connect(student_query)

if student[0]["gender"] == 'F':
    stud_address = "Ms."
else:
    stud_address = "Mr."

if student[0]["middle_name"] is not None:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["middle_name"][0].upper()}. ' \
                       f'{student[0]["last_name"].upper()}'
else:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["last_name"].upper()}'

stud_address_lastname = f'{stud_address} {student[0]["last_name"]}'

# registrar query
uni_reg_query = "SELECT * FROM registrar " \
                "WHERE registrar_position = 'University Registrar'"
uni_registrar = connect(uni_reg_query)

# receipt query
receipt_query = "SELECT receipt.or_no, " \
                "receipt.amount, receipt.date, " \
                "semester.sem, semester.acad_year FROM receipt " \
                "LEFT JOIN semester " \
                "ON receipt.sem_id = semester.sem_id " \
                "LEFT JOIN undergrad_stud " \
                "ON receipt.student_id = undergrad_stud.student_id " \
                "WHERE receipt.or_no = 1"

receipt = connect(receipt_query)

or_no = f'{receipt[0]["or_no"]:07d}'
receipt_date = custom_strftime('%m-%d-%y', receipt[0]["date"])

# dates
current_day = custom_strftime('{S}', dt.now())
current_month_year = custom_strftime('%B, %Y', dt.now())
current_date = custom_strftime('%B %d, %Y', dt.now())

# WORKBOOK

wb = Workbook()

wb.create_sheet("Page 6", 0)
ws = wb["Page 6"]

ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '14in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.75
ws.page_margins.right = 0.1
ws.page_margins.top = 0.50
ws.page_margins.bottom = 0.50

ws.column_dimensions["I"].width = 14
document_font = Font(name='Times New Roman', size=12)

# Heading
ws['B1'].value = "Republic of the Philippines"
ws.merge_cells('B1:I1')
ws['B2'].value = "Bicol University"
ws.merge_cells('B2:I2')
ws['B3'].value = "OFFICE OF THE UNIVERISTY REGISTRAR"
ws.merge_cells('B3:I3')
ws['B4'].value = "Legazpi City"
ws.merge_cells('B4:I4')
ws['B5'].value = "Tel. No. (052) 820-6809"
ws.merge_cells('B5:I5')
ws['B6'].value = "e-mail add: bu_uro@yahoo.com"
ws.merge_cells('B6:I6')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        cell.font = document_font

ws['B3'].font = Font(bold=True, name='Times New Roman')

# BU Logo
bu_logo = Image(r"static/images/bicol-university-logo.png")

bu_logo.height = 105
bu_logo.width = 105

ws.add_image(bu_logo, "A1")

ws['A7'] = 'ISO 9001:2008'
ws.merge_cells('A7:B7')
ws['A8'] = ' Certificate No.'
ws.merge_cells('A8:B8')
ws['A9'] = 'TUV100 05 1782'
ws.merge_cells('A9:B9')

iso_font = Font(color='83afdf', name='Helvetica', size=11, bold=True)
iso_alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
ws['A7'].font = iso_font
ws['A7'].alignment = iso_alignment
ws['A8'].font = iso_font
ws['A8'].alignment = iso_alignment
ws['A9'].font = iso_font
ws['A9'].alignment = iso_alignment

# Title
title_font = Font(name='Times New Roman', underline="double", size=16, bold=True)
ws['A11'].value = "HONORABLE DISMISSAL"
ws.merge_cells('A11:I11')
ws['A11'].alignment = Alignment(wrap_text=True, horizontal='center',
                               vertical='center')
ws['A11'].font = title_font

# Date
ws['I13'].value = f"{current_date}"
ws['I13'].font = document_font
ws['I13'].alignment = Alignment(horizontal='right')

# Salutation
ws['A14'].value = "TO WHOM IT MAY CONCERN"
ws['A14'].font = Font(name='Times New Roman',size=13, bold=True)

# Letter Paragraph 1
rows = ws.iter_cols(min_row=16, min_col=1, max_row=60, max_col=9)
for row in rows:
    for cell in row:
        cell.font = document_font
        
ws['A16'].value = f"This is to certify that {student_fullname}, " \
                f"who graduated with/took up subjects "
ws.merge_cells('A16:I16')
ws['A16'].font = document_font
ws['A16'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', indent=4)
ws['A17'].value = f"subjects towards the degree of {student[0]['course_name']} " \
                  f"({student[0]['course_abbrev']}) from"
ws.merge_cells('A17:I17')
ws['A18'].value = f"Bicol University {college_name}, " \
                  f"{college_address} is hereby granted honorable"
ws.merge_cells('A18:I18')
ws['A19'].value = "dismissal effective this date."
ws.merge_cells('A19:I19')
    
# Letter Paragraph 2    
ws['A20'].value = "His/Her Official Transcript of Record will be forwarded upon request by sending the lower"
ws.merge_cells('A20:I20')
ws['A20'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', indent=3.5)    
ws['A21'].value = "portion of this honorable dismissal to the college."
ws.merge_cells('A21:I21')
    
# univ registrar 
center_align = Alignment(wrap_text=True, horizontal='center', vertical='center')

ws['F24'].value = f"{uni_registrar[0]['registrar_name'].upper()}"
ws['F24'].alignment = center_align
ws.merge_cells('F24:H24')
ws['F25'].font = Font(name='Times New Roman',size=12, bold=True)
ws['F25'].value = f"{uni_registrar[0]['registrar_position'].title()}"
ws['F25'].alignment = center_align
ws.merge_cells('F25:H25')

# lower portion
ws['A26'].value = "________________________________________________________________________________"
ws.merge_cells('A26:I26')

# school name and address
rows = ws.iter_cols(min_row=27, min_col=3, max_row=30, max_col=7)
for row in rows:
    for cell in row:
        cell.alignment = center_align

ws['C27'].value = "_______________________________________"
ws.merge_cells('C27:G27')
ws['C28'].value = "____________________________________"
ws.merge_cells('C28:G28')
ws['C29'].value = "__________________________________"
ws.merge_cells('C29:G29')
ws['C30'].value = "(Complete Name of School and Address)"
ws.merge_cells('C30:G30')

# Date 2
ws['H31'].value = "_________________"
ws['H31'].alignment = center_align
ws.merge_cells('H31:I31')
ws['H32'].value = "Date"
ws['H32'].alignment = center_align
ws.merge_cells('H32:I32')

# Inside Address
red_font = Font(color="c1312c", name='Times New Roman', size=12)
ws['A33'].value = "The Registrar"
ws['A33'].font = red_font
ws.merge_cells('A33:G33')
ws['A34'].value = f"BU {college_name.title()}"
ws['A34'].font = red_font
ws.merge_cells('A34:G34')
ws['A35'].value = f"{college_address}"
ws['A35'].font = red_font
ws.merge_cells('A35:G35')

# Greetings
ws['A37'].value = "Sir/Madam:"
ws.merge_cells('A37:G37')

#body p1
indent_align = Alignment(wrap_text=True, horizontal='center', vertical='center', indent=3.5)
justify_align = Alignment(wrap_text=True, horizontal='justify')

ws['A39'].value = f"Mr/Ms {student_fullname}, " \
    f"who graduated with/took up subjects towards"
ws['A39'].alignment =  indent_align
ws.merge_cells('A39:I39')
ws['A40'].value = f"towards the degree of {student[0]['course_name']} " \
                  f"({student[0]['course_abbrev']}) from Bicol "
ws['A40'].alignment =  justify_align
ws.merge_cells('A40:I40')
ws['A41'].value = f"University {college_name}, {college_address}, " \
                  f"is temporarily enrolled in the"
ws['A41'].alignment =  justify_align
ws.merge_cells('A41:I41')
ws['A42'].value = "___________________________________ " \
                  "pending receipt of his/her Official Transcript of Record."
ws.merge_cells('A42:I42')

# body p2
ws['A44'].value = "In connection with this, may I request that his/her Official" \
                    "Transcript of Record be sent to this"
ws['A44'].alignment =  indent_align
ws.merge_cells('A44:I44')
ws['A45'].value = "University/school immediately."
ws['A45'].alignment =  justify_align
ws.merge_cells('A45:I45')

# Requesting Officer
ws['G47'].value = "_________________________"
ws['G47'].alignment = center_align
ws['G47'].font = Font(bold=True)
ws.merge_cells('G47:I47')
ws['G48'].value = "Requesting Officer"
ws['G48'].alignment = center_align
ws.merge_cells('G48:I48')
ws['G49'].value = "_________________________"
ws['G49'].font = Font(bold=True)
ws['G49'].alignment = center_align
ws.merge_cells('G49:I49')
ws['G50'].value = "Designation"
ws['G50'].alignment = center_align
ws.merge_cells('G50:I50')

# student signature
bold_italic = Font(bold=True, italic=True, name="Times New Roman")
italic_font = Font(italic=True, name="Times New Roman")

ws['A51'].value = "Not valid unless the student's signature is affixed below."
ws['A51'].font = bold_italic
ws.merge_cells('A51:F51')
ws['A52'].value = "_________________________________________________"
ws['A52'].alignment = center_align
ws['A52'].font = Font(bold=True)
ws.merge_cells('A52:E52')
ws['A53'].value = "(Signature over printed name of student)"
ws['A53'].alignment = center_align
ws['A53'].font = italic_font
ws.merge_cells('A53:E53')

# receipt
ws['A55'].value = "O.R. No."
ws['B55'].value = f"{or_no}"
ws['B55'].font = red_font
ws['C55'].value = "Amount:"
ws['C55'].alignment = Alignment(wrapText=True, horizontal='right')
ws['D55'].value = f"{receipt[0]['amount']}"
ws['D55'].font = red_font
ws['F55'].value = "Date:"
ws['F55'].alignment = Alignment(wrapText=True, horizontal='right')
ws['G55'].value = f"{receipt_date}"
ws['G55'].font = red_font

#term enrolled
ws['A56'].value = "Last Term Enrolled:"
ws.merge_cells('A56:B56')
ws['C56'].value = f"{receipt[0]['sem'][:7]}. SY {receipt[0]['acad_year']}"
ws['C56'].font = red_font
ws.merge_cells('C56:E56')

#footer
footer_font = Font(name="Arial Black", size=9)
ws['A58'].value = "BU-F-UREG-08"
ws['A58'].font = footer_font
ws['A59'].value = "Effectivity Date: Mar. 9, 2011"
ws['A59'].font = footer_font
ws['I58'].value = "Revision: 1"
ws['I58'].font = footer_font
ws['I58'].alignment = Alignment(horizontal='right')


wb.save('static/Page_6.xlsx')