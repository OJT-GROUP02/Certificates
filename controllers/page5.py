import itertools
from datetime import datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))


# db connection
def connect(query):
    cur = psycopg2.connect(database='newdb', user='postgres',
                           password='april17', host='localhost',
                           port="5432").cursor()
    cur.execute(query)
    desc = cur.description
    column_names = [col[0] for col in desc]
    result = [dict(itertools.zip_longest(column_names, row)) for row in
              cur.fetchall()]
    # result = cur.fetchall()

    return result


# QUERIES

# fetch header data
header_query = "SELECT * FROM campus_college_institute WHERE cci_id = 10"
header = connect(header_query)
college_name = header[0]["cci_name"]
college_name_upper = header[0]["cci_name"].upper()
college_address = header[0]["address"]
college_tel_no = header[0]["tel_no"]

# fetch student data
student_query = "SELECT undergrad_stud.first_name, " \
                "undergrad_stud.middle_name, undergrad_stud.last_name, " \
                "undergrad_stud.gender, undergrad_stud.date_graduated, " \
                "degree_courses.course_name, degree_courses.course_abbrev, " \
                "majors.major FROM undergrad_stud " \
                "LEFT JOIN degree_courses " \
                "ON undergrad_stud.course_id = degree_courses.course_id " \
                "LEFT JOIN majors " \
                "ON undergrad_stud.major_id = majors.major_id " \
                "WHERE undergrad_stud.student_id = 3"

student = connect(student_query)

if student[0]["gender"] == 'F':
    stud_address = "Ms."
else:
    stud_address = "Mr."

if student[0]["middle_name"] is not None:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["middle_name"][0].upper()}. ' \
                       f'{student[0]["last_name"].upper()}'
else:
    student_fullname = f'{stud_address} {student[0]["first_name"].upper()} ' \
                       f'{student[0]["last_name"].upper()}'

stud_address_lastname = f'{stud_address} {student[0]["last_name"]}'

# registrar
reg_query = "SELECT * FROM registrar " \
            "WHERE registrar_id = 2"
registrar = connect(reg_query)
uni_reg_query = "SELECT * FROM registrar " \
                "WHERE registrar_id = 1"
uni_registrar = connect(uni_reg_query)

# dates
current_day = custom_strftime('{S}', dt.now())
current_month_year = custom_strftime('%B, %Y', dt.now())
date_graduated = custom_strftime('%B %d, %Y', student[0]["date_graduated"])

# WORKBOOK
# page 5

wb = Workbook()

wb.create_sheet("Page 5", 0)
ws = wb["Page 5"]

ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.75
ws.page_margins.right = 0.75
# ws.page_margins.left = 0.75
# ws.page_margins.right = 0.1
ws.page_margins.top = 1
ws.page_margins.bottom = 1

ws.column_dimensions["I"].width = 14

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:I1')
ws.append(['Bicol University'])
ws.merge_cells('A2:I2')
ws.append([college_name_upper])
ws.merge_cells('A3:I3')
ws.append([college_address])
ws.merge_cells('A4:I4')
ws.append([college_tel_no])
ws.merge_cells('A5:I5')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=5, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')
        cell.font = Font(name='Times New Roman', size=12)

rows = ws.iter_cols(min_row=3, min_col=1, max_row=5, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(color="c1312c", name='Times New Roman', size=12)

ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                     bold=True)

# BU Logo
bu_logo = Image(r"../static/images/bicol-university-logo.png")

bu_logo.height = 105
bu_logo.width = 105

ws.add_image(bu_logo, "A1")

ws['A7'] = 'ISO 9001:2008'
ws.merge_cells('A7:B7')
ws['A8'] = ' Certificate No.'
ws.merge_cells('A8:B8')
ws['A9'] = 'TUV100 05 1782'
ws.merge_cells('A9:B9')

ws['A7'].font = Font(color='83afdf', name='Helvetica', size=11, bold=True)
ws['A7'].alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')
ws['A8'].font = Font(color='83afdf', name='Helvetica', size=11)
ws['A8'].alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')
ws['A9'].font = Font(color='83afdf', name='Helvetica', size=10, bold=True)
ws['A9'].alignment = Alignment(wrap_text=True, horizontal='left',
                               vertical='center')

# Title
ws['A11'].value = 'C E R T I F I C A T I O N'
ws.merge_cells('A11:I12')

rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
for row in rows:
    for cell in row:
        cell.font = Font(name='Times New Roman', size=20, bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                   vertical='center')

ws['A15'].value = 'TO WHOM IT MAY CONCERN:'
ws.merge_cells('A15:I15')
ws['A15'].alignment = Alignment(horizontal='left')
ws['A15'].font = Font(name='Times New Roman', size=12, bold=True)

# paragraph 1
ws['A18'].value = f"This is to certify that {student_fullname} has " \
                  f"graduated with the degree of"
ws.merge_cells('A18:I18')
ws['A18'].alignment = Alignment(horizontal='left', indent=5)
ws['A18'].font = Font(name='Times New Roman', size=12)
ws['A19'].value = f"{student[0]['course_name']} " \
                  f"({student[0]['course_abbrev']}) " \
                  f"major in {student[0]['major']} on {date_graduated} per "
ws.merge_cells('A19:I19')
ws['A20'].value = f"Resolution No. 1, s. 1997 of the Board of Regents, " \
                  f"Bicol University."

rows = ws.iter_cols(min_row=19, min_col=1, max_row=20, max_col=9)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='left')
        cell.font = Font(name='Times New Roman', size=12)

# paragraph 2
ws['A22'].value = f"It is further certified that the Bicol University is a " \
                  f"government institution operating under R.A."
ws.merge_cells('A22:I22')
ws['A22'].alignment = Alignment(horizontal='left', indent=5)
ws['A22'].font = Font(name='Times New Roman', size=12)
ws['A23'].value = f"No. 5521, as amended, whose graduates are not issued " \
                  f"Special Order"
ws.merge_cells('A23:I23')
ws['A23'].alignment = Alignment(horizontal='left')
ws['A23'].font = Font(name='Times New Roman', size=12)

# paragraph 3
ws['A25'].value = f"This certifies finally, that the " \
                  f"{student[0]['course_name']}  " \
                  f"course is duly"
ws.merge_cells('A25:I25')
ws['A25'].alignment = Alignment(horizontal='left', indent=5)
ws['A25'].font = Font(name='Times New Roman', size=12)
ws['A26'].value = f"approved by the Board of Regents, Bicol University, " \
                  f"Legazpi City."
ws.merge_cells('A26:I26')
ws['A26'].alignment = Alignment(horizontal='left')
ws['A26'].font = Font(name='Times New Roman', size=12)

# paragraph 4
ws['A28'].value = f"Issued this {current_day} day of {current_month_year} " \
                  f"upon the request of {stud_address_lastname} for " \
                  f"reference purposes."
ws.merge_cells('A28:I28')
ws['A28'].alignment = Alignment(horizontal='left', indent=5)
ws['A28'].font = Font(name='Times New Roman', size=12)

# registrar II
ws['G32'].value = f"{registrar[0]['registrar_name'].upper()}"
ws.merge_cells('G32:I32')
ws['G32'].alignment = Alignment(horizontal='center')
ws['G32'].font = Font(color='c1312c', name='Times New Roman', size=12,
                      bold=True)
ws['G33'].value = f"Registrar II"
ws.merge_cells('G33:I33')
ws['G33'].alignment = Alignment(horizontal='center')
ws['G33'].font = Font(color='c1312c', name='Times New Roman', size=12)

# noted
ws['A35'].value = "Noted:"
ws['A35'].alignment = Alignment(horizontal='left')
ws['A35'].font = Font(name='Times New Roman', size=12)
ws['B37'].value = f"{uni_registrar[0]['registrar_name'].upper()}"
ws.merge_cells('B37:D37')
ws['B37'].alignment = Alignment(horizontal='center')
ws['B37'].font = Font(name='Times New Roman', size=12, bold=True)
ws['B38'].value = "University Registrar"
ws.merge_cells('B38:D38')
ws['B38'].alignment = Alignment(horizontal='center')
ws['B38'].font = Font(name='Times New Roman', size=12)

# BU-F-REG
ws['A42'].value = 'BU-F-UREG-07'
ws.merge_cells('A42:B42')
ws['A42'].font = Font(name='Arial Black', size=9, bold=True)
ws['A43'].value = 'Effectivity Date: Mar. 9, 2011'
ws.merge_cells('A43:C43')
ws['A43'].font = Font(name='Arial Black', size=9, bold=True)
ws['I43'].value = 'Revision: 1'
ws['I43'].font = Font(name='Arial Black', size=9, bold=True)
ws['I43'].alignment = Alignment(horizontal='right')

wb.save('D:\\Users\\iveej\\Desktop\\web2py\\applications\\Certificates\\static'
        '\\Certificate_Page_5.xlsx')
