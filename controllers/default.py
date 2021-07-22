import datetime
from datetime import datetime as dt
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image


def index():
    ustudents = db().select(
        db.undergrad_stud.student_id,
        db.undergrad_stud.first_name,
        db.undergrad_stud.middle_name,
        db.undergrad_stud.last_name,
        db.undergrad_stud.gender,
        db.undergrad_stud.date_graduated,
        db.degree_courses.course_name,
        db.majors.major,
        left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                   db.degree_courses.course_id),
              db.majors.on(db.undergrad_stud.major_id == db.majors.major_id)],
        orderby=db.undergrad_stud.last_name)

    gstudents = db().select(
        db.grad_stud.student_id,
        db.grad_stud.first_name,
        db.grad_stud.middle_name,
        db.grad_stud.last_name,
        db.grad_stud.gender,
        db.degree_courses.course_name,
        db.majors.major,
        db.campus_college_institute.cci_name,
        db.campus_college_institute.address,
        db.campus_college_institute.tel_no,
        left=[db.degree_courses.on(db.grad_stud.course_id ==
                                   db.degree_courses.course_id),
              db.majors.on(db.grad_stud.major_id == db.majors.major_id),
              db.campus_college_institute.on(db.degree_courses.cci_id ==
                                             db.campus_college_institute.cci_id)])

    return locals()


def suffix(d):
    return 'th' if 11 <= d <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(d % 10,
                                                                      'th')


def custom_strftime(format_date, t):
    return t.strftime(format_date).replace('{S}', str(t.day) + suffix(t.day))


def page1():
    # get student id from URL
    try:
        stud_id = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        student = db(db.grad_stud.student_id == stud_id).select(
            db.grad_stud.first_name,
            db.grad_stud.middle_name,
            db.grad_stud.last_name,
            db.grad_stud.gender,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            db.campus_college_institute.cci_id,
            db.campus_college_institute.cci_name,
            db.campus_college_institute.address,
            db.campus_college_institute.tel_no,
            left=[db.degree_courses.on(db.grad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.grad_stud.major_id == db.majors.major_id),
                  db.campus_college_institute.on(db.degree_courses.cci_id == db.campus_college_institute.cci_id)])

        current_day = custom_strftime('{S}', dt.now())
        current_month_year = custom_strftime('%B, %Y', dt.now())
        college_name = student[0].campus_college_institute.cci_name
        college_address = student[0].campus_college_institute.address
        college_tel_no = student[0].campus_college_institute.tel_no

        if student[0].majors.major is not None:
            major = student[0].majors.major
        else:
            major = "<Insert major>"

        if student[0].grad_stud.gender == 'F':
            address = 'Ms.'
        else:
            address = 'Mr.'

        if student[0].grad_stud.middle_name is not None:
            full_name = address + ' ' + student[0].grad_stud.first_name.upper() + ' ' + \
                student[0].grad_stud.middle_name[0].upper() + '. ' + \
                student[0].grad_stud.last_name.upper()
        else:
            full_name = address + ' ' + student[0].grad_stud.first_name.upper() + ' ' + \
                student[0].grad_stud.last_name.upper()

        address_stud = address + ' ' + student[0].grad_stud.last_name

        course = student[0].degree_courses.course_name + ' (' + \
            student[0].degree_courses.course_abbrev.upper() + ')'

        registrar = db((db.registrar.college_id == student[0].campus_college_institute.cci_id)).select(
            db.registrar.registrar_name)

        registrar_II = registrar[0].registrar_name

        uni_registrar = db((db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_name)

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":
            # PAGE 1 WORKBOOK

            wb = Workbook()

            wb.create_sheet("Page1", 0)

            ws = wb["Page1"]
            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '13in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.50
            ws.page_margins.rigt = 0.50

            ws.column_dimensions["I"].width = 14

            # Heading
            ws.append(['Republic of the Philippines'])
            ws.merge_cells('A1:I1')
            ws.append(['Bicol University'])
            ws.merge_cells('A2:I2')
            ws.append([college_name.upper()])
            ws.merge_cells('A3:I3')
            ws['A3'].font = Font(bold=True)
            ws.append([college_address])
            ws.merge_cells('A4:I4')
            ws.append([college_tel_no])
            ws.merge_cells('A5:I5')

            rows = ws.iter_cols(min_row=1, min_col=1, max_row=5, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True,
                                               horizontal='center',
                                               vertical='center')
                    cell.font = Font(name='Times New Roman', size=12)

            rows = ws.iter_cols(min_row=3, min_col=1, max_row=5, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(color="c1312c", name='Times New Roman',
                                     size=12)

            ws['A3'].font = Font(color="c1312c", name='Times New Roman',
                                 size=12,
                                 bold=True)

            # BU LOGO
            bu_logo = Image(
                r"applications/Certificates/static/images/bicol-university-logo.png")

            bu_logo.height = 105
            bu_logo.width = 105

            ws.add_image(bu_logo, "A1")

            ws['A7'] = 'ISO 9001:2008'
            ws.merge_cells('A7:B7')
            ws['A8'] = ' Certificate No.'
            ws.merge_cells('A8:B8')
            ws['A9'] = 'TUV100 05 1782'
            ws.merge_cells('A9:B9')

            rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
            for row in rows:
                for cell in row:
                    cell.font = Font(color='83afdf', name='Helvetica', size=11,
                                     bold=True)
                    cell.alignment = Alignment(wrap_text=True,
                                               horizontal='left',
                                               vertical='center')

            # Title
            ws['A11'].value = "C E R T I F I C A T I O N"
            ws.merge_cells('A11:I12')

            rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(name='Times New Roman', size=20,
                                     bold=True)
                    cell.alignment = Alignment(wrap_text=True,
                                               horizontal='center',
                                               vertical='center')

            # Greetings
            ws['A15'].value = "TO WHOM IT MAY CONCERN:"
            ws.merge_cells('A15:D15')
            ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

            # Letter Body
            ws[
                'A18'].value = f"On the basis of records on file in this Office, this is " \
                               f"to certify that {full_name},"
            ws['A18'].alignment = Alignment(horizontal='left', indent=2)
            ws.merge_cells('A18:I18')
            ws[
                'A19'].value = "a graduate student of this University, has completed the Academic Requirements (CAR) for the"
            ws['A19'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A19:I19')
            ws['A20'].value = f"degree {course} major in {major}."
            ws['A20'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A20:I20')

            ws['A22'].value = f"Issued this {current_day} day of " \
                              f"{current_month_year} upon the request of " \
                              f"{address_stud} for reference"
            ws['A22'].alignment = Alignment(horizontal='left', indent=2)
            ws.merge_cells('A22:I22')
            ws['A23'].value = "purposes."
            ws['A23'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A23:I23')

            font = Font(name='Times New Roman', size=12)
            rows = ws.iter_cols(min_row=18, min_col=1, max_row=23, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = font

            # Signatories
            ws['F27'].value = f"{registrar_II.upper()}"
            ws['F27'].alignment = Alignment(horizontal='center')
            ws['F27'].font = Font(bold=True, name='Times New Roman', size=12)
            ws.merge_cells('F27:I27')
            ws['F28'].value = "Registrar II"
            ws['F28'].alignment = Alignment(horizontal='center')
            ws['F28'].font = Font(name='Times New Roman', size=12)
            ws.merge_cells('F28:I28')

            ws['A30'].value = "Noted:"
            ws['A30'].alignment = Alignment(horizontal='left')
            ws['A30'].font = Font(name='Times New Roman', size=12)
            ws.merge_cells('A30:B30')

            ws['A32'].value = f"{uni_registrar[0].registrar_name.upper()}"
            ws['A32'].alignment = Alignment(horizontal='center')
            ws['A32'].font = Font(bold=True, name='Times New Roman', size=12)
            ws.merge_cells('A32:D32')
            ws['B33'].value = "University Registrar"
            ws['B33'].alignment = Alignment(horizontal='center')
            ws['B33'].font = Font(name='Times New Roman', size=12)
            ws.merge_cells('B33:C33')

            # Footer
            ws['A40'].value = "BU-F-UREG-04"
            ws['A40'].alignment = Alignment(horizontal='left')
            ws['A40'].font = Font(bold=True, size=10)
            ws.merge_cells('A40:B40')
            ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
            ws['A41'].alignment = Alignment(horizontal='left')
            ws['A41'].font = Font(bold=True, size=10)
            ws.merge_cells('A41:C41')

            ws['H40'].value = "Revision 1"
            ws['H40'].alignment = Alignment(horizontal='right')
            ws['H40'].font = Font(bold=True, size=10)
            ws.merge_cells('H40:I40')

            current_date = custom_strftime('%Y_%m_%d', dt.now())

            if {student[0].grad_stud.first_name} is not None:
                file_name = f"{student[0].grad_stud.last_name}_{student[0].grad_stud.first_name}_{student[0].grad_stud.middle_name}_Grad_Certificate_{current_date}"
            else:
                file_name = f"{student[0].grad_stud.last_name}_{student[0].grad_stud.first_name}_Grad_Certificate_{current_date}"

            wb.save(
                f"applications/Certificates/static/documents/{file_name}.xlsx")
            raise HTTP(
                200, f"SUCCESS! Certificate for {student[0].grad_stud.first_name} {student[0].grad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")

    except IndexError:
        raise HTTP(
            400, f"HTTP 400: BAD REQUEST - Cannot process request")


def page2():
    # get student id from URL
    try:
        stud_id = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        current_day = custom_strftime('{S}', dt.now())
        current_month_year = custom_strftime('%B, %Y', dt.now())

        student = db(db.undergrad_stud.student_id == stud_id).select(
            db.undergrad_stud.first_name,
            db.undergrad_stud.middle_name,
            db.undergrad_stud.last_name,
            db.undergrad_stud.gender,
            db.undergrad_stud.date_graduated,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.undergrad_stud.major_id == db.majors.major_id)])

        if student[0].undergrad_stud.date_graduated is not None:
            date_graduated = custom_strftime('%B %d, %Y', student[
                0].undergrad_stud.date_graduated)
            year_graduated = custom_strftime('%Y', student[
                0].undergrad_stud.date_graduated)
        else:
            date_graduated = "<Insert date graduated>"
            year_graduated = "<Insert year>"

        if student[0].undergrad_stud.gender == 'F':
            address = 'Ms.'
        else:
            address = 'Mr.'

        if student[0].undergrad_stud.middle_name is not None:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.middle_name[0].upper() + '. ' + \
                student[0].undergrad_stud.last_name.upper()
        else:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.last_name.upper()

        address_stud = address + ' ' + student[0].undergrad_stud.last_name

        course = student[0].degree_courses.course_name + ' (' + \
            student[0].degree_courses.course_abbrev.upper() + ')'

        uni_registrar = db((db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_name)

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":
            # WORKBOOK
            wb = Workbook()

            wb.create_sheet("Page2", 0)

            ws = wb["Page2"]
            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '13in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.50
            ws.page_margins.rigt = 0.50

            ws.column_dimensions["I"].width = 14

            # Heading
            ws.append(['Republic of the Philippines'])
            ws.merge_cells('A1:I1')
            ws.append(['Bicol University'])
            ws.merge_cells('A2:I2')
            ws.append(['OFFICE OF THE UNIVERSITY REGISTRAR'])
            ws.merge_cells('A3:I3')
            ws['A3'].font = Font(bold=True)
            ws.append(['Legazpi City'])
            ws.merge_cells('A4:I4')
            ws.append(['Tel No. (052) 820-6809'])
            ws.merge_cells('A5:I5')
            ws.append(['E-mail Add: bu_uro@yahoo.com'])
            ws.merge_cells('A6:I6')

            rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')
                    cell.font = Font(name='Times New Roman', size=12)

            rows = ws.iter_cols(min_row=3, min_col=1, max_row=6, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(
                        color="c1312c", name='Times New Roman', size=12)

            ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                                 bold=True)

            # BU Logo
            bu_logo = Image(
                r"applications/Certificates/static/images/bicol-university-logo.png")

            bu_logo.height = 105
            bu_logo.width = 105

            ws.add_image(bu_logo, "A1")

            ws['A7'] = 'ISO 9001:2008'
            ws.merge_cells('A7:B7')
            ws['A8'] = ' Certificate No.'
            ws.merge_cells('A8:B8')
            ws['A9'] = 'TUV100 05 1782'
            ws.merge_cells('A9:B9')

            rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
            for row in rows:
                for cell in row:
                    cell.font = Font(
                        color='83afdf', name='Helvetica', size=11, bold=True)
                    cell.alignment = Alignment(wrap_text=True, horizontal='left',
                                               vertical='center')

            # Title
            ws['A11'].value = "C E R T I F I C A T I O N"
            ws.merge_cells('A11:I12')

            rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(name='Times New Roman',
                                     size=20, bold=True)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')

            # Greetings
            ws['A15'].value = "TO WHOM IT MAY CONCERN:"
            ws.merge_cells('A15:D15')
            ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

            # Letter Body
            ws['A18'].value = f"This is to certify that {full_name} has " \
                f"graduated with the degree of"
            ws['A18'].alignment = Alignment(horizontal='left', indent=5)
            ws.merge_cells('A18:I18')
            ws['A19'].value = f"{course} from Bicol University, Legazpi City on {date_graduated}"
            ws['A19'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A19:I19')
            ws['A20'].value = f"per Referendum No. 1, s, {year_graduated} of the Board of Regents."
            ws['A20'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A20:I20')

            ws['A22'].value = "It is further certified that ENGLISH LANGUAGE is the medium of instruction in all"
            ws['A22'].alignment = Alignment(horizontal='left', indent=5)
            ws.merge_cells('A22:I22')
            ws['A23'].value = "courses and in all Program Levels of Bicol University. Thus, all the academic documents such as"
            ws['A23'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A23:I23')
            ws['A24'].value = "Diploma, Transcript of Record, course descriptions, etc. are translated into English language."
            ws['A24'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A24:I24')

            ws['A26'].value = f"Issued this {current_day} day of {current_month_year} " \
                f"upon the request of {address_stud} for reference"
            ws['A26'].alignment = Alignment(horizontal='left', indent=5)
            ws.merge_cells('A26:I26')
            ws['A27'].value = "purposes."
            ws['A27'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A27:I27')

            font = Font(name='Times New Roman', size=12)
            rows = ws.iter_cols(min_row=18, min_col=1, max_row=27, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = font

            # Signatories
            ws['F31'].value = f"{uni_registrar[0].registrar_name.upper()}"
            ws['F31'].alignment = Alignment(horizontal='center')
            ws['F31'].font = Font(bold=True, name='Times New Roman', size=12)
            ws.merge_cells('F31:I31')
            ws['F32'].value = "University Registrar"
            ws['F32'].alignment = Alignment(horizontal='center')
            ws['F32'].font = Font(name='Times New Roman', size=12)
            ws.merge_cells('F32:I32')

            # Footer
            ws['A40'].value = "BU-F-UREG-04"
            ws['A40'].alignment = Alignment(horizontal='left')
            ws['A40'].font = Font(bold=True, size=10)
            ws.merge_cells('A40:B40')
            ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
            ws['A41'].alignment = Alignment(horizontal='left')
            ws['A41'].font = Font(bold=True, size=10)
            ws.merge_cells('A41:C41')

            ws['H40'].value = "Revision 1"
            ws['H40'].alignment = Alignment(horizontal='right')
            ws['H40'].font = Font(bold=True, size=10)
            ws.merge_cells('H40:I40')

            current_date = custom_strftime('%Y_%m_%d', dt.now())

            if {student[0].undergrad_stud.first_name} is not None:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_{student[0].undergrad_stud.middle_name}_Certificate01_{current_date}"
            else:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_Certificate01_{current_date}"

            wb.save(
                f"applications/Certificates/static/documents/{file_name}.xlsx")

            raise HTTP(
                200, f"SUCCESS! Certificate for {student[0].undergrad_stud.first_name} {student[0].undergrad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")
    except IndexError:
        raise HTTP(400, f"HTTP 400: BAD REQUEST - Cannot process request")


def page3():
    # get student id from URL
    try:
        stud_id = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        current_day = custom_strftime('{S}', dt.now())
        current_month_year = custom_strftime('%B, %Y', dt.now())

        student = db(db.undergrad_stud.student_id == stud_id).select(
            db.undergrad_stud.first_name,
            db.undergrad_stud.middle_name,
            db.undergrad_stud.last_name,
            db.undergrad_stud.gender,
            db.undergrad_stud.date_graduated,
            db.undergrad_stud.gwa,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            db.campus_college_institute.cci_name,
            db.campus_college_institute.address,
            db.campus_college_institute.campus,
            left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.undergrad_stud.major_id ==
                               db.majors.major_id),
                  db.campus_college_institute.on(db.degree_courses.cci_id == db.campus_college_institute.cci_id)])

        if student[0].undergrad_stud.date_graduated is not None:
            date_graduated = custom_strftime('%B %d, %Y', student[
                0].undergrad_stud.date_graduated)
            year_graduated = custom_strftime('%Y', student[
                0].undergrad_stud.date_graduated)
        else:
            date_graduated = "<Insert date graduated>"
            year_graduated = "<Insert year>"

        if student[0].undergrad_stud.middle_name is not None:
            full_name = student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.middle_name[0].upper() + '. ' + \
                student[0].undergrad_stud.last_name.upper()
        else:
            full_name = student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.last_name.upper()

        college_name = student[0].campus_college_institute.cci_name
        college_address = student[0].campus_college_institute.address
        campus = student[0].campus_college_institute.campus
        course = student[0].degree_courses.course_name + \
            ' (' + student[0].degree_courses.course_abbrev.upper() + ')'

        if student[0].majors.major is not None:
            major = student[0].majors.major
        else:
            major = "<Insert major>"

        if student[0].undergrad_stud.gwa is not None:
            gwa = student[0].undergrad_stud.gwa
        else:
            gwa = "<Insert GWA>"

        uni_registrar = db((db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_name)

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":
            # PAGE 3
            wb = Workbook()

            wb.create_sheet("Page3", 0)

            ws = wb["Page3"]
            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '13in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.50
            ws.page_margins.rigt = 0.50

            ws.column_dimensions["I"].width = 15

            # Heading
            ws.append(['Republic of the Philippines'])
            ws.merge_cells('A1:I1')
            ws.append(['Bicol University'])
            ws.merge_cells('A2:I2')
            ws.append(['OFFICE OF THE UNIVERSITY REGISTRAR'])
            ws.merge_cells('A3:I3')
            ws['A3'].font = Font(bold=True)
            ws.append(['Legazpi City'])
            ws.merge_cells('A4:I4')
            ws.append(['Tel No. (052) 820-6809'])
            ws.merge_cells('A5:I5')
            ws.append(['E-mail Add: bu_uro@yahoo.com'])
            ws.merge_cells('A6:I6')

            rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')
                    cell.font = Font(name='Times New Roman', size=12)

            rows = ws.iter_cols(min_row=3, min_col=1, max_row=6, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(
                        color="c1312c", name='Times New Roman', size=12)

            ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                                 bold=True)

            # LOGO
            bu_logo = Image(
                r"applications/Certificates/static/images/bicol-university-logo.png")

            bu_logo.height = 105
            bu_logo.width = 105

            ws.add_image(bu_logo, "A1")

            ws['A7'] = 'ISO 9001:2008'
            ws.merge_cells('A7:B7')
            ws['A8'] = 'TUV Rheinland ID 910863351'
            ws.merge_cells('A8:B8')

            rows = ws.iter_cols(min_row=7, min_col=1, max_row=9, max_col=2)
            for row in rows:
                for cell in row:
                    cell.font = Font(
                        color='83afdf', name='Helvetica', size=11, bold=True)
                    cell.alignment = Alignment(wrap_text=True, horizontal='left',
                                               vertical='center')

            # Title
            ws['A11'].value = "C E R T I F I C A T I O N"
            ws.merge_cells('A11:I12')

            rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(name='Times New Roman',
                                     size=20, bold=True)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')

            # Greetings
            ws['A15'].value = "TO WHOM IT MAY CONCERN:"
            ws.merge_cells('A15:D15')
            ws['A15'].font = Font(bold=True, name='Times New Roman', size=12)

            # Letter Body
            ws['A18'].value = f"This is to certify that {full_name} has graduated with the degree of"
            ws['A18'].alignment = Alignment(horizontal='left', indent=2)
            ws.merge_cells('A18:I18')
            ws['A19'].value = f"{course}, major in {major}"
            ws['A19'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A19:I19')
            ws['A20'].value = f"from {college_name}, {campus}, {college_address} on {date_graduated}"
            ws['A20'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A20:I20')
            ws['A21'].value = f"per Board of Regents Referendum No. 02-A, s. {year_graduated} having a General Weighted Average (GWA)"
            ws['A21'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A21:I21')
            ws['A22'].value = f"of {gwa}"
            ws['A22'].alignment = Alignment(horizontal='left')
            ws.merge_cells('A22:I22')

            ws['A24'].value = f"Issued this {current_day} day of {current_month_year} upon the request of interested party for reference purposes."
            ws['A24'].alignment = Alignment(horizontal='left', indent=2)
            ws.merge_cells('A24:I24')

            font = Font(name='Times New Roman', size=12)
            rows = ws.iter_cols(min_row=18, min_col=1, max_row=25, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = font

            # Signatories
            ws['G30'].value = f"{uni_registrar[0].registrar_name.upper()}"
            ws['G30'].alignment = Alignment(horizontal='center')
            ws['G30'].font = Font(bold=True, name='Times New Roman', size=12)
            ws.merge_cells('G30:I30')
            ws['G31'].value = f"University Registrar"
            ws['G31'].alignment = Alignment(horizontal='center')
            ws['G31'].font = Font(name='Times New Roman', size=12)
            ws.merge_cells('G31:I31')

            # Footer
            ws['A40'].value = "BU-F-UREG-05"
            ws['A40'].alignment = Alignment(horizontal='left')
            ws['A40'].font = Font(bold=True, size=10)
            ws.merge_cells('A40:B40')
            ws['A41'].value = "Effectivity Date: Mar. 9, 2011"
            ws['A41'].alignment = Alignment(horizontal='left')
            ws['A41'].font = Font(bold=True, size=10)
            ws.merge_cells('A41:C41')

            ws['H40'].value = "Revision 1"
            ws['H40'].alignment = Alignment(horizontal='right')
            ws['H40'].font = Font(bold=True, size=10)
            ws.merge_cells('H40:I40')

            current_date = custom_strftime('%Y_%m_%d', dt.now())

            if {student[0].undergrad_stud.first_name} is not None:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_{student[0].undergrad_stud.middle_name}_Certificate02_{current_date}"
            else:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_Certificate02_{current_date}"

            wb.save(
                f"applications/Certificates/static/documents/{file_name}.xlsx")

            raise HTTP(
                200, f"SUCCESS! Certificate for {student[0].undergrad_stud.first_name} {student[0].undergrad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")
    except IndexError:
        raise HTTP(400, f"HTTP 400: BAD REQUEST - Cannot process request")


def page4():
    # get student id from URL
    try:
        stud_id = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        current_day = custom_strftime('{S}', dt.now())
        current_month_year = custom_strftime('%B, %Y', dt.now())

        student = db(db.undergrad_stud.student_id == stud_id).select(
            db.undergrad_stud.first_name,
            db.undergrad_stud.middle_name,
            db.undergrad_stud.last_name,
            db.undergrad_stud.gender,
            db.undergrad_stud.date_graduated,
            db.undergrad_stud.award_id,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            db.awards.award_title,
            left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.undergrad_stud.major_id ==
                               db.majors.major_id),
                  db.awards.on(db.undergrad_stud.award_id == db.awards.award_id)])

        if student[0].undergrad_stud.date_graduated is not None:
            date_graduated = custom_strftime(
                '%B %d, %Y', student[0].undergrad_stud.date_graduated)
            year_graduated = custom_strftime('%Y', student[
                0].undergrad_stud.date_graduated)
        else:
            date_graduated = "<Insert date graduated>"
            year_graduated = "<Insert year>"

        if student[0].undergrad_stud.gender == 'F':
            address = 'Ms.'
            pronoun_1 = 'she'
            pronoun_2 = 'her'
        else:
            address = 'Mr.'
            pronoun_1 = 'he'
            pronoun_2 = 'his'

        if student[0].undergrad_stud.middle_name is not None:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + \
                ' ' + student[0].undergrad_stud.middle_name[0].upper() + '. ' + \
                student[0].undergrad_stud.last_name.upper()
        else:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.last_name.upper()

        address_stud = address + ' ' + student[0].undergrad_stud.last_name

        course = student[0].degree_courses.course_name + \
            ' (' + student[0].degree_courses.course_abbrev.upper() + ')'

        if student[0].majors.major is not None:
            major = student[0].majors.major
        else:
            major = "<Insert major>"

        if student[0].undergrad_stud.award_id is not None:
            award = student[0].awards.award_title.upper()
        else:
            award = "<Insert award>"

        uni_registrar = db((db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_name)

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":
            wb = Workbook()

            wb.create_sheet("Page 4", 0)

            ws = wb["Page 4"]
            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '13in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.50
            ws.page_margins.rigt = 0.50
            document_font = Font(name="Times New Roman", size=12)

            # column sizes
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 32
            ws.column_dimensions["C"].width = 9
            ws.column_dimensions["D"].width = 9
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 25

            # Image
            img = Image(
                r"applications/Certificates/static/images/bicol-university-logo.png")
            img.anchor = 'B1'
            img.width = 105
            img.height = 105
            ws.add_image(img)

            # Header
            ws['B1'].value = "Republic of the Philippines"
            ws.merge_cells('B1:F1')
            ws['B2'].value = 'Bicol University'
            ws.merge_cells('B2:F2')
            ws['B3'].value = 'OFFICE OF THE UNIVERSITY REGISTRAR'
            ws.merge_cells('B3:F3')
            ws['B3'].font = Font(name="Times New Roman", size=12, bold=True)
            ws['B4'].value = 'Legazpi City'
            ws.merge_cells('B4:F4')

            ws['B5'].value = 'Tel. No (052) 820-6809'
            ws.merge_cells('B5:F5')
            ws['B6'].value = 'E-mail Add: bu_uro@yahoo.com'
            ws.merge_cells('B6:F6')

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True, horizontal='center', vertical='center')
                    cell.font = document_font

            ws['B7'].value = "ISO 9001:2008"
            ws['B7'].font = Font(color="57C4E5")
            ws['B7'].alignment = Alignment(horizontal='left')

            ws['B8'].value = "Certificate No."
            ws['B8'].font = Font(color="57C4E5")
            ws['B8'].alignment = Alignment(horizontal='left')

            ws['B9'].value = "TUV100 05 1782"
            ws['B9'].font = Font(color="57C4E5")
            ws['B9'].alignment = Alignment(horizontal='left')

            ws['A12'].value = "C E R T I F I C A T I O N"
            ws.merge_cells('A12:F13')
            ws['A12'].alignment = Alignment(
                horizontal='center', vertical='center', wrapText=True)
            ws['A12'].font = Font(size='23', bold=True)

            ws['B15'].value = 'TO WHOM IT MAY CONCERN:'
            ws['B15'].font = Font(name="Times New Roman", size=12, bold=True)
            ws['A15'].alignment = Alignment(horizontal='left')

            # Body
            rows = ws.iter_cols(min_row=16, min_col=1, max_row=40, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = document_font

            # 1st Paragraph
            ws['B18'].value = f'         This is to certify that {full_name} has graduated with the degree of'
            ws.merge_cells('B18:F18')
            ws['B19'].value = f'{course} major in {major}, {award}'
            ws.merge_cells('B19:F19')
            ws['B20'].value = f' on {date_graduated} per Resolution No. 1, s. {year_graduated} of the Board of Regents, Bicol University.'
            ws.merge_cells('B20:F20')

            # 2nd Paragraph
            ws['B22'].value = f'         It is further certified that {pronoun_1} is of good moral character and has never been subjected to '
            ws.merge_cells('B23:F23')
            ws['B23'].value = f'any disciplinary action during {pronoun_2} entire stay in this University.'
            ws.merge_cells('B24:F24')

            # 3rd Paragraph
            ws['B25'].value = f'         Issued this {current_day} day of {current_month_year} upon the request of {address_stud} for reference purposes.'
            ws.merge_cells('B26:F26')

            # Registrar
            center_align = Alignment(
                wrap_text=True, horizontal='center', vertical='center')
            bold_font = Font(bold=True, name="Times New Roman", size=12)
            ws['F31'].value = uni_registrar[0].registrar_name.upper()
            ws['F31'].font = bold_font
            ws['F31'].alignment = center_align
            ws['F32'].value = 'University Registrar'
            ws['F32'].alignment = center_align

            # Footer
            footer_font = Font(name="Arial Black", size=9)
            ws['B37'].value = 'BU-F-UREG-54'
            ws['B37'].font = footer_font
            ws['F37'].value = 'Revision: 1'
            ws['F37'].font = footer_font
            ws['B38'].value = 'Effectivity Date: Mar. 9,2011'
            ws['B38'].font = footer_font

            current_date = custom_strftime('%Y_%m_%d', dt.now())

            if {student[0].undergrad_stud.first_name} is not None:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_{student[0].undergrad_stud.middle_name}_Certificate03_{current_date}"
            else:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_Certificate03_{current_date}"

            wb.save(
                f"applications/Certificates/static/documents/{file_name}.xlsx")
            raise HTTP(
                200, f"SUCCESS! Certificate for {student[0].undergrad_stud.first_name} {student[0].undergrad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")
    except IndexError:
        raise HTTP(400, f"HTTP 400: BAD REQUEST - Cannot process request")


def page5():
    # get student id from URL
    try:
        stud_arg = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        current_day = custom_strftime('{S}', dt.now())
        current_month_year = custom_strftime('%B, %Y', dt.now())

        student = db((db.undergrad_stud.student_id == stud_arg)).select(
            db.undergrad_stud.first_name,
            db.undergrad_stud.middle_name,
            db.undergrad_stud.last_name,
            db.undergrad_stud.gender,
            db.undergrad_stud.date_graduated,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            db.campus_college_institute.cci_id,
            db.campus_college_institute.cci_name,
            db.campus_college_institute.address,
            db.campus_college_institute.tel_no,
            left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.undergrad_stud.major_id ==
                               db.majors.major_id),
                  db.campus_college_institute.on(db.degree_courses.cci_id ==
                                                 db.campus_college_institute.cci_id)])

        if student[0].undergrad_stud.date_graduated is not None:
            date_graduated = custom_strftime(
                '%B %d, %Y', student[0].undergrad_stud.date_graduated)
            year_graduated = custom_strftime('%Y', student[
                0].undergrad_stud.date_graduated)
        else:
            date_graduated = "<Insert date graduated>"
            year_graduated = "<Insert year>"

        if student[0].undergrad_stud.gender == 'F':
            address = 'Ms.'
        else:
            address = 'Mr.'

        if student[0].undergrad_stud.middle_name is not None:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + \
                ' ' + student[0].undergrad_stud.middle_name[0].upper() + '. ' + \
                student[0].undergrad_stud.last_name.upper()
        else:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.last_name.upper()

        address_stud = address + ' ' + student[0].undergrad_stud.last_name

        college_name = student[0].campus_college_institute.cci_name
        college_address = student[0].campus_college_institute.address
        college_tel_no = student[0].campus_college_institute.tel_no
        course = student[0].degree_courses.course_name + \
            ' (' + student[0].degree_courses.course_abbrev.upper() + ')'

        if student[0].majors.major is not None:
            major = student[0].majors.major
        else:
            major = "<Insert major>"

        registrar = db((db.registrar.college_id == student[0].campus_college_institute.cci_id)).select(
            db.registrar.registrar_name)

        uni_registrar = db((db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_name)

        registrar_II = registrar[0].registrar_name

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":
            # WORKBOOK
            # page 5

            wb = Workbook()

            wb.create_sheet("Page 5", 0)
            ws = wb["Page 5"]

            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '13in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.75
            ws.page_margins.right = 0.75
            # ws.page_margins.left = 0.75
            # ws.page_margins.right = 0.1
            ws.page_margins.top = 1
            ws.page_margins.bottom = 1

            ws.column_dimensions["I"].width = 14

            # Heading
            ws.append(['Republic of the Philippines'])
            ws.merge_cells('A1:I1')
            ws.append(['Bicol University'])
            ws.merge_cells('A2:I2')
            ws.append([college_name.upper()])
            ws.merge_cells('A3:I3')
            ws.append([college_address])
            ws.merge_cells('A4:I4')
            ws.append([college_tel_no])
            ws.merge_cells('A5:I5')

            rows = ws.iter_cols(min_row=1, min_col=1,
                                max_row=5, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')
                    cell.font = Font(name='Times New Roman', size=12)

            rows = ws.iter_cols(min_row=3, min_col=1,
                                max_row=5, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(
                        color="c1312c", name='Times New Roman', size=12)

            ws['A3'].font = Font(color="c1312c", name='Times New Roman', size=12,
                                 bold=True)

            # BU Logo
            bu_logo = Image(
                r"applications/Certificates/static/images/bicol-university-logo.png")

            bu_logo.height = 105
            bu_logo.width = 105

            ws.add_image(bu_logo, "A1")

            ws['A7'] = 'ISO 9001:2008'
            ws.merge_cells('A7:B7')
            ws['A8'] = ' Certificate No.'
            ws.merge_cells('A8:B8')
            ws['A9'] = 'TUV100 05 1782'
            ws.merge_cells('A9:B9')

            ws['A7'].font = Font(
                color='83afdf', name='Helvetica', size=11, bold=True)
            ws['A7'].alignment = Alignment(wrap_text=True, horizontal='left',
                                           vertical='center')
            ws['A8'].font = Font(color='83afdf', name='Helvetica', size=11)
            ws['A8'].alignment = Alignment(wrap_text=True, horizontal='left',
                                           vertical='center')
            ws['A9'].font = Font(
                color='83afdf', name='Helvetica', size=10, bold=True)
            ws['A9'].alignment = Alignment(wrap_text=True, horizontal='left',
                                           vertical='center')

            # Title
            ws['A11'].value = 'C E R T I F I C A T I O N'
            ws.merge_cells('A11:I12')

            rows = ws.iter_cols(min_row=11, min_col=1, max_row=12, max_col=9)
            for row in rows:
                for cell in row:
                    cell.font = Font(name='Times New Roman',
                                     size=20, bold=True)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')

            ws['A15'].value = 'TO WHOM IT MAY CONCERN:'
            ws.merge_cells('A15:I15')
            ws['A15'].alignment = Alignment(horizontal='left')
            ws['A15'].font = Font(name='Times New Roman', size=12, bold=True)

            # paragraph 1
            ws['A18'].value = f"This is to certify that {full_name} has " \
                f"graduated with the degree of"
            ws.merge_cells('A18:I18')
            ws['A18'].alignment = Alignment(horizontal='left', indent=5)
            ws['A18'].font = Font(name='Times New Roman', size=12)
            ws['A19'].value = f"{course} major in {major} on {date_graduated} per "
            ws.merge_cells('A19:I19')
            ws['A20'].value = f"Resolution No. 1, s. {year_graduated} of the Board of Regents, Bicol University."

            rows = ws.iter_cols(min_row=19, min_col=1, max_row=20, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')
                    cell.font = Font(name='Times New Roman', size=12)

            # paragraph 2
            ws['A22'].value = f"It is further certified that the Bicol University is a " \
                f"government institution operating under R.A."
            ws.merge_cells('A22:I22')
            ws['A22'].alignment = Alignment(horizontal='left', indent=5)
            ws['A22'].font = Font(name='Times New Roman', size=12)
            ws['A23'].value = f"No. 5521, as amended, whose graduates are not issued " \
                f"Special Order"
            ws.merge_cells('A23:I23')
            ws['A23'].alignment = Alignment(horizontal='left')
            ws['A23'].font = Font(name='Times New Roman', size=12)

            # paragraph 3
            ws['A25'].value = f"This certifies finally, that the {student[0].degree_courses.course_name} course is duly"
            ws.merge_cells('A25:I25')
            ws['A25'].alignment = Alignment(horizontal='left', indent=5)
            ws['A25'].font = Font(name='Times New Roman', size=12)
            ws['A26'].value = f"approved by the Board of Regents, Bicol University, Legazpi City."
            ws.merge_cells('A26:I26')
            ws['A26'].alignment = Alignment(horizontal='left')
            ws['A26'].font = Font(name='Times New Roman', size=12)

            # paragraph 4
            ws['A28'].value = f"Issued this {current_day} day of {current_month_year} upon the request of {address_stud} for reference purposes."
            ws.merge_cells('A28:I28')
            ws['A28'].alignment = Alignment(horizontal='left', indent=5)
            ws['A28'].font = Font(name='Times New Roman', size=12)

            # registrar II
            ws['G32'].value = f"{registrar_II.upper()}"
            ws.merge_cells('G32:I32')
            ws['G32'].alignment = Alignment(horizontal='center')
            ws['G32'].font = Font(color='c1312c', name='Times New Roman', size=12,
                                  bold=True)
            ws['G33'].value = f"Registrar II"
            ws.merge_cells('G33:I33')
            ws['G33'].alignment = Alignment(horizontal='center')
            ws['G33'].font = Font(
                color='c1312c', name='Times New Roman', size=12)

            # noted
            ws['A35'].value = "Noted:"
            ws['A35'].alignment = Alignment(horizontal='left')
            ws['A35'].font = Font(name='Times New Roman', size=12)
            ws['B37'].value = f"{uni_registrar[0].registrar_name.upper()}"
            ws.merge_cells('B37:D37')
            ws['B37'].alignment = Alignment(horizontal='center')
            ws['B37'].font = Font(name='Times New Roman', size=12, bold=True)
            ws['B38'].value = "University Registrar"
            ws.merge_cells('B38:D38')
            ws['B38'].alignment = Alignment(horizontal='center')
            ws['B38'].font = Font(name='Times New Roman', size=12)

            # BU-F-REG
            ws['A42'].value = 'BU-F-UREG-07'
            ws.merge_cells('A42:B42')
            ws['A42'].font = Font(name='Arial Black', size=9, bold=True)
            ws['A43'].value = 'Effectivity Date: Mar. 9, 2011'
            ws.merge_cells('A43:C43')
            ws['A43'].font = Font(name='Arial Black', size=9, bold=True)
            ws['I43'].value = 'Revision: 1'
            ws['I43'].font = Font(name='Arial Black', size=9, bold=True)
            ws['I43'].alignment = Alignment(horizontal='right')

            current_date = custom_strftime('%Y_%m_%d', dt.now())

            if {student[0].undergrad_stud.first_name} is not None:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_{student[0].undergrad_stud.middle_name}_Certificate04_{current_date}"
            else:
                file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_Certificate04_{current_date}"

            wb.save(
                f"applications/Certificates/static/documents/{file_name}.xlsx")

            raise HTTP(
                200, f"SUCCESS! Certificate for {student[0].undergrad_stud.first_name} {student[0].undergrad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")

    except IndexError:
        raise HTTP(400, f"HTTP 400: BAD REQUEST - Cannot process request")


def page6():
    # get student id from URL
    try:
        stud_arg = int(request.args(0))
    except TypeError:
        redirect(URL('index.html'))

    try:
        current_date = str(datetime.datetime.today().strftime('%B %d, %Y'))

        student = db((db.undergrad_stud.student_id == stud_arg)).select(
            db.undergrad_stud.first_name,
            db.undergrad_stud.middle_name,
            db.undergrad_stud.last_name,
            db.undergrad_stud.gender,
            db.undergrad_stud.date_graduated,
            db.degree_courses.course_name,
            db.degree_courses.course_abbrev,
            db.majors.major,
            db.campus_college_institute.cci_name,
            db.campus_college_institute.address,
            db.campus_college_institute.tel_no,
            left=[db.degree_courses.on(db.undergrad_stud.course_id ==
                                       db.degree_courses.course_id),
                  db.majors.on(db.undergrad_stud.major_id ==
                               db.majors.major_id),
                  db.campus_college_institute.on(db.degree_courses.cci_id ==
                                                 db.campus_college_institute.cci_id)])

        receipt = db((db.receipt.student_id == stud_arg)).select(
            db.receipt.or_no,
            db.receipt.date,
            db.receipt.amount,
            db.semester.sem,
            db.semester.acad_year,
            left=[db.semester.on(db.receipt.sem_id == db.semester.sem_id),
                  db.undergrad_stud.on(db.receipt.student_id ==
                                       db.undergrad_stud.student_id)])

        if student[0].undergrad_stud.gender == 'F':
            address = 'Ms.'
        else:
            address = 'Mr.'

        if student[0].undergrad_stud.middle_name is not None:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + \
                ' ' + student[0].undergrad_stud.middle_name[0].upper() + '. ' + \
                student[0].undergrad_stud.last_name.upper()
        else:
            full_name = address + ' ' + student[0].undergrad_stud.first_name.upper() + ' ' + \
                student[0].undergrad_stud.last_name.upper()

        college_name = student[0].campus_college_institute.cci_name
        college_address = student[0].campus_college_institute.address
        course = student[0].degree_courses.course_name + \
            ' (' + student[0].degree_courses.course_abbrev.upper() + ')'

        try:
            or_no = f'{receipt[0].receipt.or_no:07d}'
            receipt_date = custom_strftime('%m-%d-%y', receipt[0].receipt.date)
            last_term_enrolled = receipt[0].semester.sem[:7] + \
                '. SY ' + receipt[0].semester.acad_year
            receipt_amount = receipt[0].receipt.amount
        except IndexError:
            or_no = '<OR no.>'
            receipt_date = '<date>'
            last_term_enrolled = '<last term enrolled>'
            receipt_amount = '<amount>'

        uni_registrar = db(
            (db.registrar.registrar_position == 'University Registrar')).select(
            db.registrar.registrar_position, db.registrar.registrar_name)

        if request.args(1) is None or request.args(1) != "download":
            return locals()
        elif request.args(1) == "download":

            # WORKBOOK
            # page 6
            wb = Workbook()

            wb.create_sheet("Page 6", 0)
            ws = wb["Page 6"]

            ColumnDimension(ws, auto_size=True)
            ws.page_setup.paperHeight = '14in'
            ws.page_setup.paperWidth = '8.5in'
            ws.page_margins.left = 0.75
            ws.page_margins.right = 0.1
            ws.page_margins.top = 0.50
            ws.page_margins.bottom = 0.50

            ws.column_dimensions["I"].width = 14
            document_font = Font(name='Times New Roman', size=12)

            # Heading
            ws['B1'].value = "Republic of the Philippines"
            ws.merge_cells('B1:I1')
            ws['B2'].value = "Bicol University"
            ws.merge_cells('B2:I2')
            ws['B3'].value = "OFFICE OF THE UNIVERISTY REGISTRAR"
            ws.merge_cells('B3:I3')
            ws['B4'].value = "Legazpi City"
            ws.merge_cells('B4:I4')
            ws['B5'].value = "Tel. No. (052) 820-6809"
            ws.merge_cells('B5:I5')
            ws['B6'].value = "e-mail add: bu_uro@yahoo.com"
            ws.merge_cells('B6:I6')

            rows = ws.iter_cols(min_row=1, min_col=1, max_row=6, max_col=9)
            for row in rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center',
                                               vertical='center')
                    cell.font = document_font

                ws['B3'].font = Font(bold=True, name='Times New Roman')

                # BU Logo
                bu_logo = Image(
                    r"applications/Certificates/static/images/bicol-university-logo.png")

                bu_logo.height = 105
                bu_logo.width = 105

                ws.add_image(bu_logo, "A1")

                ws['A7'] = 'ISO 9001:2008'
                ws.merge_cells('A7:B7')
                ws['A8'] = ' Certificate No.'
                ws.merge_cells('A8:B8')
                ws['A9'] = 'TUV100 05 1782'
                ws.merge_cells('A9:B9')

                iso_font = Font(color='83afdf', name='Helvetica',
                                size=11, bold=True)
                iso_alignment = Alignment(
                    wrap_text=True, horizontal='left', vertical='center')
                ws['A7'].font = iso_font
                ws['A7'].alignment = iso_alignment
                ws['A8'].font = iso_font
                ws['A8'].alignment = iso_alignment
                ws['A9'].font = iso_font
                ws['A9'].alignment = iso_alignment

                # Title
                title_font = Font(name='Times New Roman',
                                  underline="double", size=16, bold=True)
                ws['A11'].value = "HONORABLE DISMISSAL"
                ws.merge_cells('A11:I11')
                ws['A11'].alignment = Alignment(wrap_text=True, horizontal='center',
                                                vertical='center')
                ws['A11'].font = title_font

                # Date
                ws['I13'].value = f"{current_date}"
                ws['I13'].font = document_font
                ws['I13'].alignment = Alignment(horizontal='right')

                # Salutation
                ws['A14'].value = "TO WHOM IT MAY CONCERN"
                ws['A14'].font = Font(
                    name='Times New Roman', size=13, bold=True)

                # Letter Paragraph 1
                rows = ws.iter_cols(min_row=16, min_col=1,
                                    max_row=60, max_col=9)
                for row in rows:
                    for cell in row:
                        cell.font = document_font

                ws['A16'].value = f"This is to certify that {full_name}, who graduated with/took up subjects "
                ws.merge_cells('A16:I16')
                ws['A16'].font = document_font
                ws['A16'].alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center', indent=4)
                ws['A17'].value = f"subjects towards the degree of {course} from"
                ws.merge_cells('A17:I17')
                ws['A18'].value = f"Bicol University {college_name}, " \
                    f"{college_address} is hereby granted honorable"
                ws.merge_cells('A18:I18')
                ws['A19'].value = "dismissal effective this date."
                ws.merge_cells('A19:I19')

                # Letter Paragraph 2
                ws['A20'].value = "His/Her Official Transcript of Record will be forwarded upon request by sending the lower"
                ws.merge_cells('A20:I20')
                ws['A20'].alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center', indent=3.5)
                ws['A21'].value = "portion of this honorable dismissal to the college."
                ws.merge_cells('A21:I21')

                # univ registrar
                center_align = Alignment(
                    wrap_text=True, horizontal='center', vertical='center')

                ws['F24'].value = f"{uni_registrar[0]['registrar_name'].upper()}"
                ws['F24'].alignment = center_align
                ws.merge_cells('F24:H24')
                ws['F25'].font = Font(
                    name='Times New Roman', size=12, bold=True)
                ws['F25'].value = f"{uni_registrar[0]['registrar_position'].title()}"
                ws['F25'].alignment = center_align
                ws.merge_cells('F25:H25')

                # lower portion
                ws['A26'].value = "________________________________________________________________________________"
                ws.merge_cells('A26:I26')

                # school name and address
                rows = ws.iter_cols(min_row=27, min_col=3,
                                    max_row=30, max_col=7)
                for row in rows:
                    for cell in row:
                        cell.alignment = center_align

                ws['C27'].value = "_______________________________________"
                ws.merge_cells('C27:G27')
                ws['C28'].value = "____________________________________"
                ws.merge_cells('C28:G28')
                ws['C29'].value = "__________________________________"
                ws.merge_cells('C29:G29')
                ws['C30'].value = "(Complete Name of School and Address)"
                ws.merge_cells('C30:G30')

                # Date 2
                ws['H31'].value = "_________________"
                ws['H31'].alignment = center_align
                ws.merge_cells('H31:I31')
                ws['H32'].value = "Date"
                ws['H32'].alignment = center_align
                ws.merge_cells('H32:I32')

                # Inside Address
                red_font = Font(
                    color="c1312c", name='Times New Roman', size=12)
                ws['A33'].value = "The Registrar"
                ws['A33'].font = red_font
                ws.merge_cells('A33:G33')
                ws['A34'].value = f"BU {college_name.title()}"
                ws['A34'].font = red_font
                ws.merge_cells('A34:G34')
                ws['A35'].value = f"{college_address}"
                ws['A35'].font = red_font
                ws.merge_cells('A35:G35')

                # Greetings
                ws['A37'].value = "Sir/Madam:"
                ws.merge_cells('A37:G37')

                # body p1
                indent_align = Alignment(
                    wrap_text=True, horizontal='center', vertical='center', indent=3.5)
                justify_align = Alignment(wrap_text=True, horizontal='justify')

                ws['A39'].value = f"Mr/Ms {full_name}, " \
                    f"who graduated with/took up subjects towards"
                ws['A39'].alignment = indent_align
                ws.merge_cells('A39:I39')
                ws['A40'].value = f"towards the degree of {course} from Bicol "
                ws['A40'].alignment = justify_align
                ws.merge_cells('A40:I40')
                ws['A41'].value = f"University {college_name}, {college_address}, " \
                    f"is temporarily enrolled in the"
                ws['A41'].alignment = justify_align
                ws.merge_cells('A41:I41')
                ws['A42'].value = "___________________________________ " \
                    "pending receipt of his/her Official Transcript of Record."
                ws.merge_cells('A42:I42')

                # body p2
                ws['A44'].value = "In connection with this, may I request that his/her Official" \
                    "Transcript of Record be sent to this"
                ws['A44'].alignment = indent_align
                ws.merge_cells('A44:I44')
                ws['A45'].value = "University/school immediately."
                ws['A45'].alignment = justify_align
                ws.merge_cells('A45:I45')

                # Requesting Officer
                ws['G47'].value = "_________________________"
                ws['G47'].alignment = center_align
                ws['G47'].font = Font(bold=True)
                ws.merge_cells('G47:I47')
                ws['G48'].value = "Requesting Officer"
                ws['G48'].alignment = center_align
                ws.merge_cells('G48:I48')
                ws['G49'].value = "_________________________"
                ws['G49'].font = Font(bold=True)
                ws['G49'].alignment = center_align
                ws.merge_cells('G49:I49')
                ws['G50'].value = "Designation"
                ws['G50'].alignment = center_align
                ws.merge_cells('G50:I50')

                # student signature
                bold_italic = Font(bold=True, italic=True,
                                   name="Times New Roman")
                italic_font = Font(italic=True, name="Times New Roman")

                ws['A51'].value = "Not valid unless the student's signature is affixed below."
                ws['A51'].font = bold_italic
                ws.merge_cells('A51:F51')
                ws['A52'].value = "_________________________________________________"
                ws['A52'].alignment = center_align
                ws['A52'].font = Font(bold=True)
                ws.merge_cells('A52:E52')
                ws['A53'].value = "(Signature over printed name of student)"
                ws['A53'].alignment = center_align
                ws['A53'].font = italic_font
                ws.merge_cells('A53:E53')

                # receipt
                ws['A55'].value = "O.R. No."
                ws['B55'].value = f"{or_no}"
                ws['B55'].font = red_font
                ws['C55'].value = "Amount:"
                ws['C55'].alignment = Alignment(
                    wrapText=True, horizontal='right')
                ws['D55'].value = f"{receipt_amount}"
                ws['D55'].font = red_font
                ws['F55'].value = "Date:"
                ws['F55'].alignment = Alignment(
                    wrapText=True, horizontal='right')
                ws['G55'].value = f"{receipt_date}"
                ws['G55'].font = red_font

                # term enrolled
                ws['A56'].value = "Last Term Enrolled:"
                ws.merge_cells('A56:B56')
                ws['C56'].value = f"{last_term_enrolled}"
                ws['C56'].font = red_font
                ws.merge_cells('C56:E56')

                # footer
                footer_font = Font(name="Arial Black", size=9)
                ws['A58'].value = "BU-F-UREG-08"
                ws['A58'].font = footer_font
                ws['A59'].value = "Effectivity Date: Mar. 9, 2011"
                ws['A59'].font = footer_font
                ws['I58'].value = "Revision: 1"
                ws['I58'].font = footer_font
                ws['I58'].alignment = Alignment(horizontal='right')

                current_date = custom_strftime('%Y_%m_%d', dt.now())

                if {student[0].undergrad_stud.first_name} is not None:
                    file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_{student[0].undergrad_stud.middle_name}_Honorable_Dismissal_{current_date}"
                else:
                    file_name = f"{student[0].undergrad_stud.last_name}_{student[0].undergrad_stud.first_name}_Honorable_Dismissal_{current_date}"

                wb.save(
                    f"applications/Certificates/static/documents/{file_name}.xlsx")

                raise HTTP(
                    200, f"SUCCESS! Certificate for {student[0].undergrad_stud.first_name} {student[0].undergrad_stud.last_name} has been downloaded. path: applications/Certificates/static/documents/{file_name}.xlsx")
    except IndexError:
        raise HTTP(400, f"HTTP 400: BAD REQUEST - Cannot process request")
